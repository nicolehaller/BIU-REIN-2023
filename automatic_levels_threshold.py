import openpyxl #python library to read excel files
import math

def find_closest_number(arr, target):
    closest_num = None
    min_difference = float('inf')  
    for num in arr:
        if num is not None:
            difference = abs(num - target)
            if difference < min_difference:
                closest_num = num
                min_difference = difference
    return closest_num

#source: geeksforgeeks
def count_occurrences_to_right_from_index(arr, target, start_index): 
    count = 0
    for i in range(start_index+1, len(arr)):
        if arr[i] == target:
            count += 1
    return count

def count_occurrences_to_left_from_index(numbers, target, index):
    count = 0
    for i in range(index - 1, -1, -1):
        if numbers[i] == target:
            count += 1
    return count

#path to the original rein file with all interactions
rein_file_given = input("enter the path of a rein file: ")

# read in the corresponding excel chart file 
excel_file = input("enter the path of an excel file: ")

wb = openpyxl.load_workbook(excel_file)
sheet = wb.active

#create a temporary text file of what the user will enter manually into Jupyter notebook(the path to the new rein file they should enter and instructions)
temp_text_file = "jupyterInput.txt"
temp_text_file_writer = open(temp_text_file, "w")

# prompts the user to enter a number for the number of levels/files they want to test
levels = int(input("Enter a value for number of levels you want to test: "))

# 1) go thru excel file and add any optional interaction values to an array called optionalValuesArray
optionalValuesArray = []
for row in sheet.iter_rows():
    for cell in row:
        gene1 = str(sheet.cell(row=1, column=cell.column).value)
        gene2 = str(sheet.cell(row=cell.row, column=1).value)
        if cell.data_type == "n" and cell.value is not None: 
            numeric_value = cell.value 
        with open(rein_file_given, "r") as file:
            lines = file.readlines()
        for line in lines: 
            if f"{gene1} {gene2} positive optional" in line and numeric_value>0:
                optionalValuesArray.append(numeric_value)
            if  f"{gene1} {gene2} negative optional" in line and numeric_value<0:
                optionalValuesArray.append(numeric_value)

# 2) sort the optional values array in ascending order
optionalValuesArray.sort()
print("optionalValuesArray: " + str(optionalValuesArray))

# 3) create an array of the dif levels of numbers of interactions we are looking for
# ex: if 50 optional interactions and 5 levels then levelsArray is [10,20,30,40,50] corresponding to number of optional interactions we want in every file, round down
levelsArray = []
numberOfOptionalValues= len(optionalValuesArray)
for i in range(0, levels):
    levelsArray.append(math.floor(len(optionalValuesArray) - (i*(numberOfOptionalValues/levels))))
print(levelsArray)

# create arrays and set their values to be None:
# threshold array - will store the threshold that yielded the number of optional interactions we are looking for (we know how many optional interactions by looking at the samen index but in the levelsArray)
threshold= [None] * len(levelsArray)
# extraArray - stores how many optional interactions would be in the file if the threshold was the threshold in the same index of the threshold array 
extraArray = [None]  * len(optionalValuesArray)

#  4) iterate through the optionalValuesArray:  
for i in range(len(optionalValuesArray)):
   indexLeftOfItsNegative = 0
    #if positive number add up right of it and left of its negative
   if optionalValuesArray[i] > 0:
       numbersToTheRight = (len(optionalValuesArray)-1) - i
       if i != len(optionalValuesArray)-1:
         if optionalValuesArray[i] == optionalValuesArray[i+1]:
            subtraction_amount = count_occurrences_to_right_from_index(optionalValuesArray, optionalValuesArray[i], i)
            numbersToTheRight = numbersToTheRight - subtraction_amount
       for j in range(len(optionalValuesArray)):
            if optionalValuesArray[j] >= -(optionalValuesArray[i]):
                indexLeftOfItsNegative= j
                break
       numbersToTheLeft = j 
   #if negative number add up left of it and right of its positive
   if optionalValuesArray[i] < 0:
        indexRightOfAbsVal = len(optionalValuesArray) #start index at rightmost element by default
        for j in range(len(optionalValuesArray)):
            if optionalValuesArray[j] > abs(optionalValuesArray[i]) and i!=j:
                indexRightOfAbsVal= j
                break
        numbersToTheRight = len(optionalValuesArray) - indexRightOfAbsVal 
        numbersToTheLeft = i
        if i != 0:
           if optionalValuesArray[i] == optionalValuesArray[i-1]:
              subtraction_amount = count_occurrences_to_left_from_index(optionalValuesArray, optionalValuesArray[i], i)
              numbersToTheLeft = numbersToTheLeft - subtraction_amount
   numbersToRightAndLeft = numbersToTheRight + numbersToTheLeft
   #if the number of optional interactions is one of the level numbers we want, add this threshold to the threshold array
   if numbersToRightAndLeft in levelsArray:
         index_found = levelsArray.index(numbersToRightAndLeft)
         threshold[index_found] = optionalValuesArray[i]
   #otherwise if the number of optional interactions isnt one of the level numbers we want, add the number of optional interactions it has to the extraArray(for later if one of the levels isnt found we want to see which is the closest number of optional interactions to the original that we could use)
   elif numbersToRightAndLeft not in levelsArray:
         index_of_threshold = i
         extraArray[index_of_threshold] = numbersToRightAndLeft
#manually set the first file's threshold to the biggest negative so that it includes all interactions
threshold[0] = -.0000000000000000001

print("extraArray: all extra stuff, none only for ones that fit the levels" + str(extraArray))

print("threshold array: " + str(threshold))

#make sure we found thresholds to output every level in the levelsArray
#if not, search the extra array for the closest number of optional interactions to the original and make that the threshold
for i in range(len(threshold)):
   if threshold[i] is None:
      ClosestLeftRightToLevel = find_closest_number(extraArray, levelsArray[i])
      closest_index = extraArray.index(ClosestLeftRightToLevel)
      threshold[i] = optionalValuesArray[closest_index]

print("new threshold array: " + str(threshold))
      
# use the thresholds from the threshold array to create levels # of rein and text files
for i in range(0, levels, 1): #0, levels, 1
   #set threshold based on which level we r on
   threshold_number = threshold[i]
   #create a rein output file called outputFile{i} that will store only optional interactions within the threshold
   #NEED::: new_rein_file = f"/Users/lilyp/Documents/submission version 6/submission/artifact/Examples/Science2014/outputFile{i}.rein"
   new_rein_file = f"outputFile{i}.rein"

   rein_file_writer = open(new_rein_file, "w")

   #add the instructions for the current file to the jupyter notebook text file
   temp_text_file_writer.write("./Science2014/" + f"outputFile{i}.rein" + "\nlet model792 = ReinAPI.LoadFile ”./Science2014/" + f"outputFile{i}.rein" + "\n Model792 |> ReinAPI.DrawBespokeNetworkWithSizeSVG 10000.0\n ReinAPI.CheckAndPrint model792" + "\n" + "the threshold was: " + str(threshold_number) + "\n"  + "\n")

   # Open the given rein input file in read mode
   with open(rein_file_given, "r") as file:
      lines = file.readlines()

   # write to the new rein file 3 parts: 
   # 1) the section above the list of gene interactions 
   # 2) the list of gene interactions but only the optional relationships within the threshold
   # 3) the section after the gene interactions (experiments and functions)

   # section 1: 
   #"positive" and "negative" are words we want to stop printing the first section at when we get to them
   stop_words = ["positive", "negative"]
   for line in lines:
      if any(word in line for word in stop_words): #once get to first line that has gene1 gene2 interaction_type(pos or neg) we will stop writing this first section to the output file
         break 
      rein_file_writer.write(line + "\n") 

   # section 2: iterate through the excel file to narrow down interactions
   for row in sheet.iter_rows():
      for cell in row:
         gene1 = str(sheet.cell(row=1, column=cell.column).value)
         gene2 = str(sheet.cell(row=cell.row, column=1).value)
         if cell.data_type == "n" and cell.value is not None: 
            numeric_value = cell.value 
            #if threshold is a positive number then print to file all relationships above that number and optional 
            if threshold_number > 0: 
                if numeric_value > threshold_number: 
                  for line in lines:
                     if f"{gene1} {gene2} positive optional" in line:
                        rein_file_writer.write(line + "\n")
                        print (line + "\n") 
                if numeric_value < -(threshold_number):
                     if f"{gene1} {gene2} negative optional" in line:
                        rein_file_writer.write(line + "\n")
                        print (line + "\n")
            #if user enters a negative number then print to file all relationships below that number and above the abs value and optional 
            else:
               if numeric_value < threshold_number or numeric_value > abs(threshold_number):
                  for line in lines:
                     if numeric_value < 0: 
                        if f"{gene1} {gene2} negative optional" in line:
                           rein_file_writer.write(line + "\n")
                           print (line + "\n")
                     else: 
                        if f"{gene1} {gene2} positive optional" in line:
                           rein_file_writer.write(line + "\n")
                           print (line + "\n")

   for line in lines:
      if ("negative" in line or "positive" in line) and not "optional" in line:
         rein_file_writer.write(line + "\n")
   # Section 3
   last_stop_word_index = -1
   # Find the index of the last line containing a stop word
   for i, line in enumerate(lines):
      if any(word in line for word in stop_words):
         last_stop_word_index = i
   rein_file_writer.write("\n".join(lines[last_stop_word_index + 1:])) #writes everything from the line after the last gene1 gene2 interaction_type line until the end of the file

temp_text_file_writer.close()
wb.close()
